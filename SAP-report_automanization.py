import datetime
import os

import openpyxl
import pandas as pd


texts = ['Общее количество / 总数',
		 'Количестиво по критичности "Очень высокая" / "很着急"危机问题的数量',
		 'Количестиво по критичности "Высокая" / "着急"危机问题的数量',
		 'Количестиво по критичности "Средняя" / "中"危机问题的数量',
		 'Количестиво по критичности "Низкая" / "低"危机问题的数量', ]


def formatting(sheet):
	sheet.column_dimensions["A"].width = 7.29  	#1
	sheet.column_dimensions["B"].width = 20.71 	#2
	sheet.column_dimensions["C"].width = 15.71 	#3
	sheet.column_dimensions["D"].width = 15.43 	#4
	sheet.column_dimensions["E"].width = 21.57 	#5
	sheet.column_dimensions["F"].width = 81    	#6
	sheet.column_dimensions["G"].width = 37.71	#7
												#8 h=критичность
												#9 i=приоритет
	sheet.column_dimensions["J"].width = 19.86	#10
	sheet.column_dimensions["K"].width = 17.71	#11
	sheet.column_dimensions["L"].width = 15.14	#12
	sheet.column_dimensions["M"].width = 18		#13

	for j in range(1, (sheet.max_column+1)):
		if j == sheet.max_column:
			sheet.cell(row=2, column=j).value = 'Кол-во дней отсутствия решения (на {})'.format(datetime.datetime.today().strftime("%d.%m.%Y"))
	for i in range(2, (sheet.max_row+1)):
		for j in range(1, (sheet.max_column+1)):
			if j == 2 or j == 12:
				sheet.cell(row=i, column=j).number_format = 'DD.MM.YYYY'
			if j == sheet.max_column:
				try:
					sheet.cell(row=(i+2), column=j).value = "{}".format(int((datetime.datetime.now()-sheet.cell(row=(i+2), column=2).value).days))
					sheet.cell(row=(i+2), column=j).number_format = '###'
				except Exception:
					continue


def summary_filling (sheet, list):
	SubTotRowNum=[]
	for k in range(4):  #СР ЗНАЧ ПО ПОДИТОГАМ
		SUM = 0
		el = 0
		for i in range(0, ((list[k][4]+2+list[k][0]+1)-(list[k][4]+3))):
			SUM += int(sheet.cell(row=((list[k][4]+3+i)), column=13).value)
			el += 1
		sheet.cell(row=(list[k][4]+2), column=13).value = (SUM//el)
		SubTotRowNum.append(list[k][4]+2)

	sheet.cell(row=3, column=13).value = 0
	for i in range(0, len(SubTotRowNum)):
		sheet.cell(row=3, column=13).value += int(sheet.cell(row=SubTotRowNum[i], column=13).value)
	sheet.cell(row=3, column=13).value = (int(sheet.cell(row=3, column=13).value) // len(SubTotRowNum))


def show_all_rows(sheet):
	for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
		if sheet.row_dimensions[row[0].row].hidden:
			sheet.row_dimensions[row[0].row].hidden = False


def remove_closed_issues(sheet):
	ClosIss = 0
	while ClosIss:
		ClosIss = 0
		for i in range(1, sheet.max_row):
			if sheet.cell(row=i, column=10).value == 'Закрыт/任务关闭':
				ClosIss += 1
				sheet.delete_rows(i)
		if ClosIss == 0:
			ClosIss = None


def sort_and_get_name(file):
	df = pd.read_excel(file)
	df['Дата фиксации требования'] = pd.to_datetime(df['Дата фиксации требования'])
	df['计划完成日期 Planned Finish'] = pd.to_datetime(df['计划完成日期 Planned Finish']) 
	names = df.columns
	df = df.sort_values([names[8], names[11], names[1], names[0]], ascending=[True, True, True, True])
	tempName = "SAP2.xlsx"
	df.to_excel(tempName, index=False)
	return tempName


def result_by_criticality (sheet):
	"""
	CritSubTot (Criticality subtotals) Подытоги по критичности:
	[Общее кол-во задач по критичности,
	 Кол-во зарегистрированных задач,
	 Кол-во задач в работе,
	 Номер строки последней задачи,
	 Строка подытогов по критичности].

	"""
	statistic = []
	CritSubTot = [0, 0, 0, 0, 0]
	for krit in range(1, 5):
		for i in range(1, sheet.max_row+1):
			if sheet.cell(row=i, column=9).value == krit:
				CritSubTot[0] += 1
				if sheet.cell(row=i, column=10).value == 'Зарегистрировано/未开始任务':
					CritSubTot[1] += 1
				if sheet.cell(row=i, column=10).value == 'В работе/任务处理中':
					CritSubTot[2] += 1
				if sheet.cell(row=(i+1), column=9).value != krit:
					CritSubTot[3] = sheet.cell(row=(i+1), column=9).row
					CritSubTot[4]=CritSubTot[3]-CritSubTot[0]+krit-1
		statistic.append(CritSubTot)
		CritSubTot = [0, 0, 0, 0, 0]
	return statistic


def create_result_rows(sheet, list, texts):
	# Создание строки общих итогов
	sheet.insert_rows(1)
	sheet.insert_rows(3)
	sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
	sheet.cell(row=3, column=1).value=texts[0]
	sheet.cell(row=3, column=3).value=0
	sheet.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
	sheet.cell(row=3, column=4).value='Зарегистрировано/未开始任务'
	sheet.cell(row=3, column=6).value=0
	sheet.cell(row=3, column=7).value='В работе/任务处理中'
	sheet.cell(row=3, column=10).value=0
	sheet.merge_cells(start_row=3, start_column=11, end_row=3, end_column=12)
	sheet.cell(row=3, column=11).value='Среднее время нахождения в работе / 在工作中平均时间'

	ALL=0
	REG=0
	INWORK=0
	for i in range(4):
		ALL+=list[i][0]
		REG +=list[i][1]
		INWORK +=list[i][2]	

	# Cоздание строк с подитогами по критичности (сколько задач: всего, зарегистрировано, в работе - общее время нахождения в работе)
	for i in range(4):
		n=list[i][4]
		sheet.insert_rows(n+2)
		sheet.merge_cells(start_row=(n+2), start_column=1, end_row=(n+2), end_column=2)
		sheet.cell(row=(n+2), column=1).value=texts[i+1]
		sheet.cell(row=(n+2), column=3).value=list[i][0]
		sheet.merge_cells(start_row=(n+2), start_column=4, end_row=(n+2), end_column=5)
		sheet.cell(row=(n+2), column=4).value='Зарегистрировано/未开始任务'
		sheet.cell(row=(n+2), column=6).value=list[i][1]
		sheet.cell(row=(n+2), column=7).value='В работе/任务处理中'
		sheet.cell(row=(n+2), column=10).value=list[i][2]
		sheet.merge_cells(start_row=(n+2), start_column=11, end_row=(n+2), end_column=12)
		sheet.cell(row=(n+2), column=11).value='Среднее время нахождения в работе / 在工作中平均时间'

	sheet.cell(row=3, column=3).value=ALL
	sheet.cell(row=3, column=6).value=REG
	sheet.cell(row=3, column=10).value=INWORK

	# Создание шапки таблицы
	sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
	todayDate = datetime.date.today().strftime("%d.%m.%Y")
	sheet.cell(row=1, column=1).value = 'Список проблем SAP на {} / SAP issues list on {}'.format(todayDate, todayDate)


def del_empty_rows(sheet):
	nstr=0
	while nstr:
		nstr=0
		for i in range(1, sheet.max_row):
			if sheet.cell(row=i, column=1).value == None:
				nstr += 1
				sheet.delete_rows(i)
		if nstr == 0:
			nstr = None


def sign_rows(sheet):
	text = ['ФД - FD', 'РФЦ - RDC', 'Дата-日期']
	for r in range(1,3):
		n=sheet.max_row
		sheet.insert_rows(n)
		sheet.merge_cells(start_row=n, start_column=1, end_row=n, end_column=2)
		sheet.cell(row=n, column=1).value= text[r-1]
		sheet.merge_cells(start_row=(n), start_column=3, end_row=n, end_column=10)
		sheet.cell(row=n, column=11).value= text[2]
		sheet.merge_cells(start_row=n, start_column=12, end_row=n, end_column=13)
		for j in range(1, sheet.max_column+1):
			sheet.cell(row=n, column=j).font = openpyxl.styles.Font(
				name='Times New Roman',
				size=12, bold=True,
				color = "000000"
			)
			sheet.cell(row=n, column=j).alignment = openpyxl.styles.Alignment(
				horizontal='center',
				vertical='center',
				wrap_text=True
			)
			sheet.cell(row=n, column=j).border = openpyxl.styles.Border(
				left=openpyxl.styles.Side(style='thin'),
				right=openpyxl.styles.Side(style='thin'),
				top=openpyxl.styles.Side(style='thin'),
				bottom=openpyxl.styles.Side(style='thin'),
			)
			sheet.row_dimensions[n].height = 80
	sheet.delete_rows(sheet.max_row)


def hide_cols(sheet):
	for col in ['H', 'I']:
		sheet.column_dimensions[col].hidden= True


def standardization(sheet,list):
	PIT=[]
	viravnivanie = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
	granici = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

	simpleFont = openpyxl.styles.Font(name='Times New Roman', size=12)
	simpleFill = openpyxl.styles.PatternFill(patternType='solid', start_color='ffffff', end_color='ffffff')
	
	TFont = openpyxl.styles.Font(name='Times New Roman', size=24, bold=True, color = "000000")

	P_IFont = openpyxl.styles.Font(name='Times New Roman', size=12, bold=True, color = "000000")
	IFill = openpyxl.styles.PatternFill(patternType='solid', start_color='F5C243', end_color='F5C243')

	PIFill = openpyxl.styles.PatternFill(patternType='solid', start_color='FEFF55', end_color='FEFF55')

	whiteFont = openpyxl.styles.Font(name='Times New Roman', size=12, bold=True, color = "ffffff")
	greenFill = openpyxl.styles.PatternFill(patternType='solid', start_color='4EAE5B', end_color='4EAE5B')
	
	for i in range(0, len(list)):
		PIT.append(list[i][4]+2)
	for i in range (1, sheet.max_row+1):
		for j in range (1, sheet.max_column+1):
			if (i == 1):
				sheet.cell(row=i, column=j).font = TFont
				sheet.cell(row=i, column=j).fill = simpleFill
				sheet.cell(row=i, column=j).alignment = viravnivanie
				sheet.cell(row=i, column=j).border = granici
				sheet.row_dimensions[i].height = 70
			
			elif (i == 2):
				sheet.cell(row=i, column=j).font = whiteFont
				sheet.cell(row=i, column=j).fill = greenFill
				sheet.cell(row=i, column=j).alignment = viravnivanie
				sheet.cell(row=i, column=j).border = granici
				sheet.row_dimensions[i].height = 70
			
			elif (i == 3):
				sheet.cell(row=i, column=j).font = P_IFont
				sheet.cell(row=i, column=j).fill = IFill
				sheet.cell(row=i, column=j).alignment = viravnivanie
				sheet.cell(row=i, column=j).border = granici
				sheet.row_dimensions[i].height = 70
			
			elif i in PIT:
				sheet.cell(row=i, column=j).font = P_IFont
				sheet.cell(row=i, column=j).fill = PIFill
				sheet.cell(row=i, column=j).alignment = viravnivanie
				sheet.cell(row=i, column=j).border = granici
				sheet.row_dimensions[i].height = 70

			else:
				sheet.cell(row=i, column=j).font = simpleFont
				sheet.cell(row=i, column=j).fill = simpleFill
				sheet.cell(row=i, column=j).alignment = viravnivanie
				sheet.cell(row=i, column=j).border = granici


def str_into_date(sheet):
	for i in range(1, sheet.max_row+1):
		if (sheet.cell(row=i, column=12).value == "Дата неизвестна"):
			sheet.cell(row=i, column=12).value = "31.12.2023"
		elif (sheet.cell(row=i, column=12).value == "Артур обновит дату после уточнения деталей"):
			sheet.cell(row=i, column=12).value = "31.12.2023"
		elif (sheet.cell(row=i, column=12).value == None):
			sheet.cell(row=i, column=12).value = "31.12.2023"
		else:
			continue


def date_into_str(sheet):
	for i in range(1, sheet.max_row):
		try:
			if (sheet.cell(row=i, column=12).value.strftime('%d.%m.%Y') == "31.12.2023"):
				sheet.cell(row=i, column=12).value = "Дата неизвестна"
		except Exception:
			continue


def main():
	originalName = str(input ("Введите имя файла (без .xlsx) :\nInput file's name (without .xlsx):\n"))
	book = openpyxl.load_workbook(originalName+".xlsx")

	# Удаление колонок, которые не отображаются в отчёте
	book.active = 0 
	reportPage = book.active
	reportPage.delete_cols (idx = 19)
	reportPage.delete_cols (idx = 16, amount = 2)
	reportPage.delete_cols (idx = 12, amount = 3)

	#Раскрытие всех строк, удаление решённых задач, удаление пустых строк (бывают в конце)
	show_all_rows(reportPage)
	remove_closed_issues(reportPage)
	del_empty_rows(reportPage)

	# Преобразование текста в колонке дедлайна в дату(31.12.2023)
	str_into_date(reportPage)

	#Создание копии и сортировка записей в таблице по фильтру (Приоритет: Важность > Дедлайн > Дата создания задачи)
	todayDate = datetime.date.today().strftime("%d.%m.%Y")
	file='Список проблем SAP на {} - SAP issues list on {}.xlsx'.format(todayDate, todayDate)
	book.save(file)
	tempFile= sort_and_get_name(file)
	book = openpyxl.load_workbook(tempFile)
	reportPage = book.active

	#Cоздание строк итогов и подытогов, оформление таблицы к стандарту
	statList = result_by_criticality(reportPage)
	create_result_rows(reportPage, statList, texts)
	date_into_str(reportPage)
	formatting(reportPage)
	summary_filling(reportPage, statList)
	del_empty_rows(reportPage)
	standardization(reportPage, statList)
	sign_rows(reportPage)

	# Скрытие колонок H и I (Критичсноть, Приоритет решения)
	hide_cols(reportPage)

	#удаление временного файла
	os.remove(tempFile)

	book.active.title = "Month report"
	book.save(file)


if __name__== "__main__":
	main()
